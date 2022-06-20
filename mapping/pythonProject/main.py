import folium

from openpyxl import load_workbook

load_wb = load_workbook("C:\\Users\\zzim2\\Documents\\카카오톡 받은 파일\\대상파일.xlsx",data_only = True)

load_ws = load_wb['서울특별시_구로구_도로파손접수_20210831']

map_osm = folium.Map(location=[37.4962,126.8624],zoom_start=14)

print(str(load_ws.cell(2,6).value))

i = 1
j = 0
while True:
    i = i + 1

    if ((load_ws.cell(i, 6).value== None) and (load_ws.cell(i, 7).value == None)):
        pass
    else:
        x = float(load_ws.cell(i, 6).value)
        y = float(load_ws.cell(i, 7).value)
        detail = (load_ws.cell(i,1).value)
        print(x,y)
        folium.Marker(location = [x,y],
               popup = (detail)).add_to(map_osm)
        print(str(i)+"등록")

    if ((load_ws.cell(i,1).value) == None):
        break

    map_osm.save("C:\\Users\\zzim2\\Documents\\카카오톡 받은 파일\\map.html")





